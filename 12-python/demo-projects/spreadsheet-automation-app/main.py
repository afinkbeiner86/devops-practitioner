import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_inventory_value = {}
products_less_10 = {}

for product_row in range(2, product_list.max_row + 1):

    product_number = int(product_list.cell(product_row, 1).value)
    product_amount = int(product_list.cell(product_row, 2).value)
    product_price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    total_product_value = product_list.cell(product_row, 5)

    # Calculcate products per supplier        
    products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name, 0) + 1

    # Calculcate total inventory value per supplier
    total_inventory_value[supplier_name] = total_inventory_value.get(supplier_name, 0) + product_price * product_amount

    # Print product with amount less than 10
    if product_amount < 10:
        products_less_10[product_number] = products_less_10.get(product_number, 0) + product_amount
        print(f"\nProduct {product_number} is almost out of stock!\nAmount left: {product_amount}\n")
    
    # Calculate and write total product value in new row
    total_product_value.value = product_amount * product_price

# Save changes to new file
 #inv_file.save("inventory_with_total_value_row.xlsx")

print(f"\n---\nProducts per supplier:\n{products_per_supplier}\n---")
print(f"Total inventory value per supplier:\n{total_inventory_value}\n---")
print(f"Product's stock less than 10:\n{products_less_10}\n---")