"""
EXERCISE 9: Working with Spreadsheets
Write a program that:

reads the provided spreadsheet file "employees.xlsx" with the following information/columns: "name", "years of experience", "job title", "date of birth"
creates a new spreadsheet file "employees_sorted.xlsx" with following info/columns: "name", "years of experience", 
where the years of experience is sorted in descending order: so the employee name with the most experience in years is on top.
"""

import openpyxl

employee_file = openpyxl.load_workbook("employees.xlsx")

employee_list = employee_file["Sheet1"]

# Remove columns 3 and 4
employee_list.delete_cols(3,4)

employees_by_experience = []

# Iterate through each row and write cell 1 & 2 to a dictionary
# Append the dictionary to a list
for row in range(2, employee_list.max_row + 1):
    employee_name = employee_list.cell(row, 1).value
    employee_experience = int(employee_list.cell(row, 2).value)

    employees_by_experience.append({
        "name": employee_name,
        "experience": employee_experience
    })

# Use a lambda function to sort the list from highest to lowest
# using the value of the dictionary key "experience"
# Copy the sorted dictionaries to a new list
sorted_keys = sorted(employees_by_experience, key=lambda x: x["experience"], reverse=True)
sorted_employees_by_experience = sorted_keys.copy()

# Interate through each row and write the sorted values from the list
# of dictionaries to cell 1 (name) and 2 (experience)
index = 0
for row in range(2, employee_list.max_row + 1):
    employee_name = employee_list.cell(row, 1)
    employee_experience = employee_list.cell(row, 2)
    
    employee_name.value = sorted_employees_by_experience[index].get("name")
    employee_experience.value = sorted_employees_by_experience[index].get("experience")
    index += 1

employee_file.save("sorted_employees.xlsx")